#import tensorflow as tf
#print(tf.__version__)

import os
import cv2
import numpy as np
import random

# Set the path to the directory containing your raw face images
data_dir = "dataset"
output_dir = "preprocessed_dataset"

# Create the output directory if it doesn't exist
if not os.path.exists(output_dir):
    os.makedirs(output_dir)

# Define the image size you want for preprocessing
target_size = (128, 128)

# Define the number of augmented images per original image
num_augmentations = 6  # Adjust this number as needed

# Loop through the raw dataset and apply preprocessing steps
for root, dirs, files in os.walk(data_dir):
    for file in files:
        if file.endswith(".jpg"):
            img_path = os.path.join(root, file)

            # Load the original image
            img = cv2.imread(img_path)

            # Resize the image to the target size
            img = cv2.resize(img, target_size)

            # Convert the image to grayscale (if needed)
            img = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

            # Normalize pixel values to [0, 1]
            img = img.astype(np.float32) / 255.0

            # Perform data augmentation
            augmented_images = []
            for _ in range(num_augmentations):
                augmented_img = img.copy()

                # Apply random augmentations
                if random.choice([True, False]):
                    augmented_img = cv2.flip(augmented_img, 1)  # Horizontal flip

                if random.choice([True, False]):
                    rotation_angle = random.uniform(-10, 10)
                    rotation_matrix = cv2.getRotationMatrix2D((target_size[0] / 2, target_size[1] / 2), rotation_angle, 1.0)
                    augmented_img = cv2.warpAffine(augmented_img, rotation_matrix, target_size, flags=cv2.INTER_LINEAR)

                # You can add more augmentations as needed

                augmented_images.append(augmented_img)

            # Save the preprocessed images (including augmented versions)
            for i, augmented_img in enumerate(augmented_images):
                output_path = os.path.join(output_dir, f"{file.replace('.jpg', f'_aug{i}.jpg')}")
                cv2.imwrite(output_path, augmented_img)

print("Data preprocessing and data augmentation complete.")

from sklearn.model_selection import train_test_split
import os
import shutil

# Set the path to your master dataset directory
dataset_dir = "dataset"  # Change this to the directory where you combined the images for each person

# Create directories for the training, validation, and testing sets
train_dir = "train"
validation_dir = "validation"
test_dir = "test"
os.makedirs(train_dir, exist_ok=True)
os.makedirs(validation_dir, exist_ok=True)
os.makedirs(test_dir, exist_ok=True)

# Specify the split ratios (e.g., 60% training, 20% validation, 20% testing)
train_ratio = 0.6
validation_ratio = 0.2
test_ratio = 0.2

# Get the list of image directories (one for each person)
person_subdirectories = [d for d in os.listdir(dataset_dir) if os.path.isdir(os.path.join(dataset_dir, d))]

# Loop through the directories and split each person's images
for person_dir in person_subdirectories:
    # Get the list of image files for the current person
    image_files = [f for f in os.listdir(os.path.join(dataset_dir, person_dir)) if f.endswith(".jpg")]

    # Split the dataset for the current person
    train_files, temp_files = train_test_split(image_files, test_size=1 - train_ratio, random_state=42)
    validation_files, test_files = train_test_split(temp_files, test_size=test_ratio / (test_ratio + validation_ratio), random_state=42)

    # Create subdirectories for the current person in the train, validation, and test sets
    train_person_dir = os.path.join(train_dir, person_dir)
    validation_person_dir = os.path.join(validation_dir, person_dir)
    test_person_dir = os.path.join(test_dir, person_dir)
    os.makedirs(train_person_dir, exist_ok=True)
    os.makedirs(validation_person_dir, exist_ok=True)
    os.makedirs(test_person_dir, exist_ok=True)

    # Move the files to their respective directories
    for file in train_files:
        src = os.path.join(dataset_dir, person_dir, file)
        dst = os.path.join(train_person_dir, file)
        shutil.move(src, dst)

    for file in validation_files:
        src = os.path.join(dataset_dir, person_dir, file)
        dst = os.path.join(validation_person_dir, file)
        shutil.move(src, dst)

    for file in test_files:
        src = os.path.join(dataset_dir, person_dir, file)
        dst = os.path.join(test_person_dir, file)
        shutil.move(src, dst)

print("Data splitting complete.")

import cv2
import os
import numpy as np

# Define the directory paths for your preprocessed data
train_dir = "train"
validation_dir = "validation"
test_dir = "test"

# Define the target size for resizing the images
target_size = (128, 128)


# Function to load and preprocess images
def load_and_preprocess_images(directory):
    images = []
    labels = []

    for label, person_dir in enumerate(os.listdir(directory)):
        person_path = os.path.join(directory, person_dir)

        for image_file in os.listdir(person_path):
            image_path = os.path.join(person_path, image_file)

            # Load and preprocess the image
            image = cv2.imread(image_path)
            image = cv2.resize(image, target_size)
            image = image.astype(np.float32) / 255.0  # Normalize pixel values to [0, 1]

            images.append(image)
            labels.append(label)

    return np.array(images), np.array(labels)


# Load and preprocess training data
train_images, train_labels = load_and_preprocess_images(train_dir)

# Load and preprocess validation data
validation_images, validation_labels = load_and_preprocess_images(validation_dir)

# Load and preprocess testing data
test_images, test_labels = load_and_preprocess_images(test_dir)

# Print the shape of the datasets
print("Training images shape:", train_images.shape)
print("Training labels shape:", train_labels.shape)
print("Validation images shape:", validation_images.shape)
print("Validation labels shape:", validation_labels.shape)
print("Testing images shape:", test_images.shape)
print("Testing labels shape:", test_labels.shape)

import tensorflow as tf
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Flatten, Dense, Dropout

num_classes = 4
# Load a pre-trained VGG-16 model without the top classification layers
pretrained_model = tf.keras.applications.VGG16(weights='imagenet', include_top=False, input_shape=(128, 128, 3))

# Create a new top classification head
model = Sequential([
    pretrained_model,
    Flatten(),
    Dense(512, activation='relu'),
    Dropout(0.5),
    Dense(512, activation='relu'),
    Dropout(0.5),
    Dense(num_classes, activation='softmax')  # num_classes is the number of individuals you want to recognize
])

# Freeze the weights of the pre-trained layers
for layer in pretrained_model.layers:
    layer.trainable = False

# Compile the model
model.compile(optimizer='adam',
              loss='sparse_categorical_crossentropy',  # Use 'categorical_crossentropy' if you one-hot encode labels
              metrics=['accuracy'])

# Fine-tune the model on your dataset
history = model.fit(train_images, train_labels, epochs=10, validation_data=(validation_images, validation_labels))

# Evaluate the fine-tuned model on the test set
test_loss, test_accuracy = model.evaluate(test_images, test_labels)
print("Test accuracy:", test_accuracy)


# Save the trained model
model.save('my_face_recognition_model.h5')

import cv2
import openpyxl
import numpy as np
import tensorflow as tf
import datetime
from mtcnn.mtcnn import MTCNN  # Import MTCNN for face detection

# Define the name mapping dictionary
name_mapping = {
    0: "Nikhila",
    1: "Ashritha",
    2: "Aruna",
    3: "Deekshitha",
    4: "Ganesh"
}

# Define the department mapping dictionary
department_mapping = {
    "Nikhila": "CSE",
    "Ashritha": "CSE",
    "Aruna": "CSE",
    "Deekshitha": "CSE",
    "Ganesh": "CSE",
}

# Load the pre-trained face recognition model (adjust this)
model = tf.keras.models.load_model('my_face_recognition_model.h5')

# Load the names used in training, testing, and validation datasets
training_names = ["Nikhila", "Ashritha", "Aruna", "Deekshitha", "Ganesh"]
testing_names = ["Nikhila", "Ashritha", "Aruna", "Deekshitha", "Ganesh"]  # Add testing names if available
validation_names = ["Nikhila", "Ashritha", "Aruna", "Deekshitha", "Ganesh"]  # Add validation names if available

# Open an existing Excel workbook or create a new one
workbook = openpyxl.load_workbook("attendance.xlsx")
sheet = workbook.active

# Clear existing data in the Excel sheet, including headers
sheet.delete_rows(2, sheet.max_row - 1)

# Add headers for the "Name," "Day," "Time," "Department," and "Status" columns
sheet.cell(row=1, column=1, value="Name")
sheet.cell(row=1, column=2, value="Day")
sheet.cell(row=1, column=3, value="Time")
sheet.cell(row=1, column=4, value="Department")
sheet.cell(row=1, column=5, value="Status")

# Create a set to keep track of recognized names in the current session
recognized_names = set()
recognized_names_in_session = set()

# Set a confidence threshold for recognizing known faces
confidence_threshold = 0.7

# Initialize the video capture
video_capture = cv2.VideoCapture(0)

# Initialize MTCNN for face detection
detector = MTCNN()

while True:
    ret, img = video_capture.read()

    # Detect faces using MTCNN
    faces = detector.detect_faces(img)

    for result in faces:
        x, y, w, h = result['box']
        x, y, w, h = int(x), int(y), int(w), int(h)  # Ensure integer values
        detected_face = img[y:y + h, x:x + w]

        # Preprocess the detected face
        face = cv2.resize(detected_face, (128, 128))
        face = face.astype(np.float32) / 255.0

        # Perform face recognition
        face = np.expand_dims(face, axis=0)
        prediction = model.predict(face)
        max_confidence = np.max(prediction)
        recognized_label = np.argmax(prediction)

        if max_confidence < confidence_threshold:
            predicted_name = "Unknown"
            department = "N/A"
            status = "Absent"
        else:
            predicted_name = name_mapping.get(recognized_label, "Unknown")
            department = department_mapping.get(predicted_name, "N/A")
            status = "Present"

        if predicted_name != "Unknown" and predicted_name not in recognized_names_in_session:
            # Add the name, day, time, department, and status to the Excel sheet at the beginning
            day = datetime.datetime.now().strftime("%Y-%m-%d")
            time = datetime.datetime.now().strftime("%H:%M:%S")
            new_row = [predicted_name, day, time, department, status]
            sheet.insert_rows(2)  # Insert a new row at the beginning
            for idx, value in enumerate(new_row, start=1):
                sheet.cell(row=2, column=idx, value=value)  # Set values for each cell in the new row
            recognized_names_in_session.add(predicted_name)

        # Draw a rectangle around the detected face
        cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)  # Draw a green rectangle
        cv2.putText(img, predicted_name, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 255, 0), 1, cv2.LINE_AA)

    cv2.imshow("Face Detection", img)

    key = cv2.waitKey(1)

    if key == ord('q'):
        # Update the set of recognized names for the current session
        recognized_names.update(recognized_names_in_session)
        recognized_names_in_session.clear()

    if key == ord('x'):
        break

# Save the updated Excel sheet with recognized faces and attendance data
workbook.save("attendance.xlsx")

# Release the video capture and close all OpenCV windows
video_capture.release()
cv2.destroyAllWindows()
