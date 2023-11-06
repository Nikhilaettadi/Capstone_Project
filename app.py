from flask import Flask, render_template, Response, request, send_file, redirect, url_for, session
import cv2
import openpyxl
import numpy as np
import tensorflow as tf
import datetime
from mtcnn.mtcnn import MTCNN

app = Flask(__name__)

app.secret_key = 'Nikhilaettadi'

# Define the name mapping dictionary
name_mapping = {
    0: "Nikhila",
    1: "Ashritha",
    2: "Aruna",
    3: "Deekshitha",
    4: "Ganesh"
}

# Define the department mapping dictionary
department_mapping = {
    "Nikhila": "CSE",
    "Ashritha": "CSE",
    "Aruna": "CSE",
    "Deekshitha": "CSE",
    "Ganesh": "CSE",
}

# Load the pre-trained face recognition model (adjust this)
model = tf.keras.models.load_model('my_face_recognition_model.h5')

# Load the names used in training, testing, and validation datasets
training_names = ["Nikhila", "Ashritha", "Aruna", "Deekshitha", "Ganesh"]
testing_names = ["Nikhila", "Ashritha", "Aruna", "Deekshitha", "Ganesh"]  # Add testing names if available
validation_names = ["Nikhila", "Ashritha", "Aruna", "Deekshitha", "Ganesh"]  # Add validation names if available

# Open an existing Excel workbook or create a new one
workbook = openpyxl.load_workbook("attendance.xlsx")
sheet = workbook.active

# Clear existing data in the Excel sheet, including headers
sheet.delete_rows(2, sheet.max_row - 1)

# Add headers for the "Name," "Day," "Time," "Department," and "Status" columns
sheet.cell(row=1, column=1, value="Name")
sheet.cell(row=1, column=2, value="Day")
sheet.cell(row=1, column=3, value="Time")
sheet.cell(row=1, column=4, value="Department")
sheet.cell(row=1, column=5, value="Status")

# Create a set to keep track of recognized names in the current session
recognized_names = set()
recognized_names_in_session = set()

# Set a confidence threshold for recognizing known faces
confidence_threshold = 0.7

# Initialize the video capture (initially stopped)
video_capture = None

# Initialize MTCNN for face detection
detector = MTCNN()

# Global variable to stop video capture
stop_video_capture = False

# A simple user database (you can replace this with your own user management system)
users = {
    'Nikhila' : 'sru123'
}
# Define a route for the login page
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if users.get(username) == password:
            # If the login is successful, set a session variable and then redirect to the main page
            session['logged_in'] = True  # You need to import 'session' from 'flask'
            return redirect(url_for('index'))
        else:
            # If login fails, display an error message or redirect to the login page again
            return "Login failed. Please try again"

    return render_template('login.html')

@app.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))  # Redirect to the login page if not logged in
    return render_template('index.html')

def generate_frames():
    while True:
        if stop_video_capture:
            break  # Exit the loop

        if video_capture is not None:
            ret, img = video_capture.read()

            # Detect faces using MTCNN
            faces = detector.detect_faces(img)

            for result in faces:
                x, y, w, h = result['box']
                x, y, w, h = int(x), int(y), int(w), int(h)  # Ensure integer values
                detected_face = img[y:y + h, x:x + w]

                # Preprocess the detected face
                face = cv2.resize(detected_face, (128, 128))
                face = face.astype(np.float32) / 255.0

                # Perform face recognition
                face = np.expand_dims(face, axis=0)
                prediction = model.predict(face)
                max_confidence = np.max(prediction)
                recognized_label = np.argmax(prediction)

                if max_confidence < confidence_threshold:
                    predicted_name = "Unknown"
                    department = "N/A"
                    status = "Absent"
                else:
                    predicted_name = name_mapping.get(recognized_label, "Unknown")
                    department = department_mapping.get(predicted_name, "N/A")
                    status = "Present"

                if predicted_name != "Unknown" and predicted_name not in recognized_names_in_session:
                    # Add the name, day, time, department, and status to the Excel sheet at the beginning
                    day = datetime.datetime.now().strftime("%Y-%m-%d")
                    time = datetime.datetime.now().strftime("%H:%M:%S")
                    new_row = [predicted_name, day, time, department, status]
                    sheet.insert_rows(2)  # Insert a new row at the beginning
                    for idx, value in enumerate(new_row, start=1):
                        sheet.cell(row=2, column=idx, value=value)  # Set values for each cell in the new row
                    recognized_names_in_session.add(predicted_name)

                ret, buffer = cv2.imencode('.jpg', img)
                frame = buffer.tobytes()
                yield (b'--frame\r\n'
                    b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/start_video')
def start_video():
    global video_capture
    video_capture = cv2.VideoCapture(0)  # Start video capture
    return 'Video started'

@app.route('/stop_video')
def stop_video():
    global stop_video_capture, video_capture
    stop_video_capture = True
    if video_capture is not None:
        video_capture.release()
    cv2.destroyAllWindows()
    return 'Video stopped'

# Route for downloading the updated Excel sheet
@app.route('/download_attendance', methods=['GET', 'POST'])
def download_attendance():
    if request.method == 'POST':
        try:
            workbook.save("attendance.xlsx")  # Save the updated workbook
            return send_file("attendance.xlsx", as_attachment=True, download_name="attendance.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as e:
            return f'Error: {str(e)}'
    return render_template('index.html')

if __name__ == '__main':
    app.run(debug=True)
